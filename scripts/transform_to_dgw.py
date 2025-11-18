import os
import pandas as pd
import yaml
from openpyxl import load_workbook
from collections import defaultdict

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
CONFIG_DIR = os.path.join(BASE_DIR, "config", "mappings")
INCOMING_DIR = os.path.join(BASE_DIR, "data", "incoming")
TEMPLATES_DIR = os.path.join(BASE_DIR, "data", "templates_dgw")
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "curated")


def load_yaml(path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def detect_mapping_file(filename: str) -> str:
    """
    Decide qual arquivo de mapping (YAML) usar com base no nome do arquivo legado.
    """
    name = filename.lower()
    if "hire" in name or "hirestack" in name:
        return "mapping_hire.yaml"
    elif "contact" in name:
        return "mapping_contact.yaml"
    elif "worker" in name:
        return "mapping_worker.yaml"
    elif "absence" in name or "abs_" in name:
        return "mapping_absence.yaml"
    elif "compensation" in name or "comp_" in name:
        return "mapping_compensation.yaml"
    else:
        return "mapping_generic.yaml"


def detect_template_file(filename: str, template_files: list[str]) -> str | None:
    """
    Escolhe o template DGW correto com base no nome do arquivo legado.
    Regras simples de substring para casar tipos (Hire, Absence, Contact etc.).
    """
    name = filename.lower()

    # pares (palavra-chave no legado, palavras-chave esperadas no template)
    rules = [
        (["hirestack", "hire"], ["hirestack", "hire"]),
        (["absence", "abs_"], ["absence", "abs_"]),
        (["contact"], ["contact"]),
        (["worker"], ["worker"]),
        (["compensation", "comp_"], ["compensation", "comp_"]),
    ]

    for legacy_keys, template_keys in rules:
        if any(k in name for k in legacy_keys):
            for t in template_files:
                tl = t.lower()
                if any(k in tl for k in template_keys):
                    return t

    # fallback: tenta casar pelo prefixo antes do primeiro "_"
    prefix = name.split("_")[0]
    for t in template_files:
        if t.lower().startswith(prefix):
            return t

    # último fallback: se só tiver 1 template, usa ele
    if len(template_files) == 1:
        return template_files[0]

    return None


def get_valid_sheets(path):
    xl = pd.ExcelFile(path)
    sheets = [s for s in xl.sheet_names if not s.strip().startswith(">")]
    xl.close()
    return sheets


def transform_to_dgw():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    incoming_files = [f for f in os.listdir(INCOMING_DIR)
                      if f.lower().endswith(".xlsx")]
    template_files = [f for f in os.listdir(TEMPLATES_DIR)
                      if f.lower().endswith(".xlsx")]

    if not incoming_files:
        print("⚠️ No source files found in /data/incoming.")
        return
    if not template_files:
        print("⚠️ No DGW templates were found in /data/templates_dgw.")
        return

    print("🧩 Starting legacy template transformation...\n")

    for file in incoming_files:
        print(f"➡️ Converting {file}...")

        input_path = os.path.join(INCOMING_DIR, file)

        # 1) mapping YAML
        mapping_file = detect_mapping_file(file)
        mapping_path = os.path.join(CONFIG_DIR, mapping_file)
        if not os.path.exists(mapping_path):
            print(f"❌ Mapping not found: {mapping_file}")
            continue

        yaml_data = load_yaml(mapping_path)
        aliases = yaml_data.get("aliases", {})

        # 2) template DGW específico para este arquivo
        template_name = detect_template_file(file, template_files)
        if not template_name:
            print("❌ Could not find a matching DGW template for this file.")
            continue

        template_path = os.path.join(TEMPLATES_DIR, template_name)
        print(f"   📄 Template loaded: {template_name}")
        print(f"   📑 Mapping YAML:   {mapping_file}")

        # copia workbook do template para saída
        tmpl_wb = load_workbook(template_path)
        output_path = os.path.join(
            OUTPUT_DIR,
            f"{file.replace('.xlsx', '')}_DGW_ready.xlsx"
        )
        tmpl_wb.save(output_path)
        out_wb = load_workbook(output_path)

        # abas origem x template
        src_sheets = get_valid_sheets(input_path)
        tmpl_sheets = get_valid_sheets(template_path)

        for sheet in tmpl_sheets:
            ws = out_wb[sheet]

            if sheet not in src_sheets:
                print(f"⚠️ Sheet '{sheet}' does not exist in the legacy file — it will be left empty.")
                continue

            print(f"   📝 Filling sheet: {sheet}")

            # header na linha 2 do legado
            src_df = pd.read_excel(
                input_path,
                sheet_name=sheet,
                header=1,
                keep_default_na=False,
                na_values=[]
            )

            src_df.columns = src_df.columns.astype(str).str.strip()
            src_df = src_df.fillna("")

            # cabeçalhos do template (linha 6)
            header_row = 6
            template_headers = [
                (cell.value.strip() if isinstance(cell.value, str) else cell.value)
                for cell in ws[header_row]
            ]

            # header template -> lista de posições (para duplicados)
            header_positions = defaultdict(list)
            for col_index, header in enumerate(template_headers):
                if header:
                    header_positions[header].append(col_index)

            start_row = 7

            # percorre cada linha da origem
            for r_index, (_, row) in enumerate(src_df.iterrows()):
                if row.isna().all():
                    continue

                excel_row = start_row + r_index

                # para cada coluna do template (alias key)
                for tgt_col, alias_list in aliases.items():
                    # alias_list sempre é lista
                    if isinstance(alias_list, str):
                        alias_list = [alias_list]

                    # encontra primeiro header de origem que exista
                    src_col_name = None
                    for alias in alias_list:
                        if alias in src_df.columns:
                            src_col_name = alias
                            break

                    if not src_col_name:
                        continue  # nenhum alias presente na origem

                    if tgt_col not in header_positions:
                        continue  # template não tem esta coluna

                    value = row[src_col_name]

                    # escreve em TODAS as colunas do template com esse header
                    for col_idx in header_positions[tgt_col]:
                        ws.cell(row=excel_row, column=col_idx + 1, value=value)

        out_wb.save(output_path)
        print(f"✅ DGW file ready: {output_path}\n")

    print("✅ Transformation completed!")


if __name__ == "__main__":
    transform_to_dgw()
